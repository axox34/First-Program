"""
First Program circa July 2018

Look in Outlook email for particular files sent by customer. Download the files and process according to needs by Sales Team
Used pyinstaller to create exe but possibly had some deprecation issues after release to Sales.
1: search emails
2: download the text file
3: copy the values in the text file into an excel file
4: make slight changes to excel file for presentation
"""


from win32com.client import Dispatch
import os.path
import xlsxwriter
from datetime import datetime

import openpyxl
#back to openpyxl 2.0.2 which is in EMAUIL!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
from openpyxl.styles import Style
#------------------------------------------------------------
from openpyxl.styles import Color, Font, colors

from openpyxl.styles import PatternFill

#get date and time 
current_time = datetime.now()
date = (current_time.strftime('%m%d%y'))

#redo excel such that the numbers are floats instead of decimals
def validate_number(s):
    try:
        return float(s)
    except (ValueError, TypeError):
        return s
    
    
#convert txt into xlsx with same name
def converttotxt(file,location):
    

    global newxlsxfile
    newxlsxfile = os.path.splitext(location+ date)[0]+'.xlsx'
    workbook = xlsxwriter.Workbook(newxlsxfile)        #create file
    worksheet = workbook.add_worksheet()           #create worksheet
    data = open(file,'r')                #loaddata

    linelist = data.readlines() 

    global maxrow             #read each line
    maxrow = len(linelist)                 #count lines
    print (maxrow)                       #check number of lines

    for num in range (0, maxrow):         #create each line and print in excel

        line = linelist[num] 
        global splitline           #load each line in variable
        splitline = line.split("\t")          #split lines\
        splitline = [int(s) if s.isdigit() else s for s in splitline]
        splitline = [validate_number(s) for s in splitline]
        worksheet.write_row(num, 0, splitline)         #write each line in excel

    return workbook, worksheet, newxlsxfile, maxrow
        

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    print (rangeSelected)
    return rangeSelected


#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j ).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1



def createData(startCol, startRow, endCol, endRow):
    selectedRange = copyRange(startCol, startRow, endCol, endRow,sheet) #Change the 4 number values
    pastingRange = pasteRange(startCol, startRow, endCol, endRow,temp_sheet,selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    print (pastingRange)
    template.save(location+ date+'.xlsx')
    print("Range copied and pasted!")


#create pink highlight at total
def highlighter(temp_sheet):
    cell = temp_sheet.cell('D11')
    cell.number_format = 'General'

    count = 1
    while count <= maxrow:
            y = 'AD' + str(count)
            print(y)
            cell = temp_sheet.cell(y)
            cell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('FFFFCCE5')))
            count += 1

#access outlook
outlook = Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder("6")
all_inbox = inbox.Items


#fetch all messages, then look for the particular email
for msg in all_inbox:
    if msg.Unread and msg.SenderEmailType == "EX":
        print(msg)
        sendersemail = msg.Sender.GetExchangeUser().PrimarySmtpAddress
        if sendersemail == ("jwong@fujikin.com"): 
            if "PLAN_SCHEDULE" in msg.Subject:
                print("yes")
                for att in msg.Attachments:
                    att.SaveASFile(os.getcwd() + '\\' + att.FileName)
                    print("forecast")
                    if "SINGAPORE" in msg.Subject:
                        location = "Singapore "
                        formula = "Formula- Singapore Item List.xlsx"
                    elif "317-MRP" in msg.Subject:
                        location = "Santa Clara "
                        formula = "Formula- Santa Clara Item List.xlsx"
                    elif "SCMRP318" in msg.Subject:
                        location = "Santa Clara-02 "
                        formula = "Formula- 788 Item List.xlsx"
                    else:
                        break
                print(location)
                converttotxt("Forecast.txt",location)
                #remove txt file after the xlsx has been created
                os.remove("Forecast.txt")
                #create files to copy into
                forecastwb = newxlsxfile
                wb= openpyxl.load_workbook(forecastwb) #Add file name
                sheet = wb['Sheet1'] #Add Sheet name
                
                template = openpyxl.load_workbook(formula) #Add file name
                temp_sheet = template["Sheet1"] #Add Sheet name++
                        #adding highlights
                highlighter(temp_sheet)
                #copying data into xlsx
                createData(1,1,1,sheet.max_row)
                createData(2,1,35,10)
                
                createData(4,9,35,sheet.max_row)

           






